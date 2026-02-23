#!/usr/bin/env python3
"""
CX 대시보드 Phase 2: 세그먼트 매칭 + HTML 주입

사용법:
  python3 scripts/phase2_enrich_inject.py

동작:
  1. data/new_records.json 로드 (Phase 1 산출물)
  2. BigQuery에서 세그먼트 데이터 조회 (사업자번호 매칭)
  3. 기존 ALL_RECORDS + 신규 레코드 합치기
  4. index.html, cs-retag-dashboard.html 양쪽 업데이트
"""

import sys
import os
import json
import re
import subprocess

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
NEW_RECORDS_JSON = os.path.join(ROOT_DIR, "data", "new_records.json")
INDEX_HTML = os.path.join(ROOT_DIR, "index.html")

# ── 한도 구간 매핑 ──
def to_limit_tier(grant_limit):
    """grant_limit(원) → 한도 구간 문자열"""
    if grant_limit is None:
        return "미매칭"
    try:
        v = float(grant_limit)
    except (ValueError, TypeError):
        return "미매칭"
    if v < 10_000_000:
        return "1천만 미만"
    elif v < 50_000_000:
        return "1천만~5천만"
    elif v < 100_000_000:
        return "5천만~1억"
    elif v < 500_000_000:
        return "1억~5억"
    else:
        return "5억 이상"


def load_new_records():
    """Phase 1 산출물 로드"""
    if not os.path.exists(NEW_RECORDS_JSON):
        print(f"❌ {os.path.relpath(NEW_RECORDS_JSON, ROOT_DIR)} 없음. Phase 1을 먼저 실행하세요.")
        sys.exit(1)
    with open(NEW_RECORDS_JSON, "r", encoding="utf-8") as f:
        records = json.load(f)
    return records


def load_existing_records(html_path):
    """HTML에서 기존 ALL_RECORDS 추출"""
    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()
    match = re.search(r"const ALL_RECORDS = (\[.*?\]);", content, re.DOTALL)
    if not match:
        return [], content
    records = json.loads(match.group(1))
    return records, content


def build_corp_number_map(excel_path):
    """Excel User data에서 userId → corp_number(숫자만) 매핑"""
    try:
        import openpyxl
    except ImportError:
        print("  ⚠️  openpyxl 없음, 사업자번호 매핑 스킵")
        return {}

    # 가장 최근 Excel 파일 찾기
    if not excel_path:
        data_dir = os.path.join(ROOT_DIR, "data")
        xlsx_files = sorted(
            [f for f in os.listdir(data_dir) if f.endswith(".xlsx")],
            key=lambda f: os.path.getmtime(os.path.join(data_dir, f)),
            reverse=True,
        )
        if not xlsx_files:
            print("  ⚠️  data/ 폴더에 Excel 파일 없음")
            return {}
        excel_path = os.path.join(data_dir, xlsx_files[0])
        print(f"  Excel: {os.path.basename(excel_path)}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # UserChat에서 chatId → userId 매핑
    ws_chat = wb["UserChat"]
    chat_to_user = {}
    chat_to_corp_direct = {}
    for r in range(2, ws_chat.max_row + 1):
        chat_id = ws_chat.cell(r, 1).value
        if not chat_id:
            continue
        user_id = ws_chat.cell(r, 14).value
        if user_id:
            chat_to_user[chat_id] = user_id
        # data_only로 직접 읽히는 사업자번호도 시도
        corp_direct = ws_chat.cell(r, 4).value
        if corp_direct and isinstance(corp_direct, str):
            chat_to_corp_direct[chat_id] = re.sub(r"[^0-9]", "", corp_direct)

    # User data에서 userId → corp_number 매핑
    ws_user = wb["User data"]
    user_to_corp = {}
    for r in range(2, ws_user.max_row + 1):
        uid = ws_user.cell(r, 1).value
        corp = ws_user.cell(r, 6).value  # profile.corp_number
        if uid and corp:
            user_to_corp[uid] = re.sub(r"[^0-9]", "", str(corp))

    # chatId → corp_number(숫자만) 최종 매핑
    result = {}
    for chat_id in chat_to_user:
        # 직접 읽힌 사업자번호 우선
        if chat_id in chat_to_corp_direct:
            result[chat_id] = chat_to_corp_direct[chat_id]
        else:
            user_id = chat_to_user[chat_id]
            if user_id in user_to_corp:
                result[chat_id] = user_to_corp[user_id]

    return result


def load_segment_cache():
    """data/segment_cache.json에서 세그먼트 캐시 로드"""
    cache_path = os.path.join(ROOT_DIR, "data", "segment_cache.json")
    if not os.path.exists(cache_path):
        return {}
    with open(cache_path, "r", encoding="utf-8") as f:
        rows = json.load(f)
    seg_map = {}
    for row in rows:
        cid = str(row.get("corp_id", ""))
        seg_map[cid] = {
            "detailCategory": row.get("detail_category") or "미매칭",
            "stabilityStatus": row.get("stability_status") or "미매칭",
            "growthStatus": row.get("growth_status") or "미매칭",
            "limitTier": to_limit_tier(row.get("grant_limit")),
        }
    return seg_map


def fetch_segments_from_bq(corp_ids):
    """BigQuery에서 세그먼트 데이터 조회 (bq CLI → 캐시 폴백)"""
    if not corp_ids:
        return {}

    # corp_id 리스트를 SQL IN절로
    id_list = ", ".join(str(cid) for cid in corp_ids)
    query = f"""
    SELECT corp_id, detail_category, stability_status, growth_status, grant_limit
    FROM `gowid-prd.mart_customer_segment.segment_base`
    WHERE month_id = (SELECT MAX(month_id) FROM `gowid-prd.mart_customer_segment.segment_base`)
      AND corp_id IN ({id_list})
    """

    # bq CLI 시도
    cmd = [
        "bq", "query",
        "--use_legacy_sql=false",
        "--format=json",
        "--max_rows=10000",
        query,
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0:
            rows = json.loads(result.stdout)
            seg_map = {}
            for row in rows:
                cid = str(row.get("corp_id", ""))
                seg_map[cid] = {
                    "detailCategory": row.get("detail_category") or "미매칭",
                    "stabilityStatus": row.get("stability_status") or "미매칭",
                    "growthStatus": row.get("growth_status") or "미매칭",
                    "limitTier": to_limit_tier(row.get("grant_limit")),
                }
            # 성공 시 캐시 업데이트
            cache_path = os.path.join(ROOT_DIR, "data", "segment_cache.json")
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(rows, f, ensure_ascii=False, indent=2)
            print(f"  BQ 쿼리 성공 → 캐시 업데이트")
            return seg_map
    except (FileNotFoundError, subprocess.TimeoutExpired, json.JSONDecodeError):
        pass

    # 폴백: 캐시 파일
    print(f"  BQ 쿼리 실패 → 캐시(segment_cache.json) 사용")
    return load_segment_cache()


def enrich_records(new_records, corp_map, seg_map):
    """신규 레코드에 세그먼트 데이터 주입"""
    matched = 0
    for rec in new_records:
        chat_id = rec["chatId"]
        corp_id = corp_map.get(chat_id)
        if corp_id and corp_id in seg_map:
            seg = seg_map[corp_id]
            rec["detailCategory"] = seg["detailCategory"]
            rec["stabilityStatus"] = seg["stabilityStatus"]
            rec["growthStatus"] = seg["growthStatus"]
            rec["limitTier"] = seg["limitTier"]
            matched += 1
    return matched


def inject_into_html(html_path, all_records, fields_to_keep):
    """HTML 파일의 ALL_RECORDS를 교체"""
    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()

    # 필드 필터링
    filtered = []
    for r in all_records:
        filtered.append({k: r.get(k) for k in fields_to_keep})

    # JSON 직렬화 (한 줄)
    json_str = json.dumps(filtered, ensure_ascii=False, separators=(",", ":"))
    new_line = f"    const ALL_RECORDS = {json_str};"

    # 교체
    content = re.sub(
        r"    const ALL_RECORDS = \[.*?\];",
        new_line,
        content,
        count=1,
        flags=re.DOTALL,
    )

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(content)


def main():
    print("🚀 Phase 2: 세그먼트 매칭 + HTML 주입\n")

    # 1. 신규 레코드 로드
    print("[1/5] 신규 레코드 로드")
    new_records = load_new_records()
    print(f"  신규: {len(new_records)}건")

    if not new_records:
        print("  새 레코드 없음. 종료.")
        return

    # 2. 사업자번호 매핑 빌드
    print("[2/5] 사업자번호 매핑")
    corp_map = build_corp_number_map(None)  # 가장 최근 Excel 자동 탐색
    mapped = sum(1 for r in new_records if r["chatId"] in corp_map)
    print(f"  사업자번호 매핑: {mapped}/{len(new_records)}건")

    # 3. BQ 세그먼트 조회
    print("[3/5] BigQuery 세그먼트 조회")
    unique_corps = set(corp_map.get(r["chatId"]) for r in new_records if r["chatId"] in corp_map)
    unique_corps.discard(None)
    print(f"  고유 사업자번호: {len(unique_corps)}개")

    seg_map = fetch_segments_from_bq([int(c) for c in unique_corps if c.isdigit()])
    print(f"  BQ 매칭: {len(seg_map)}개 법인")

    # 4. 세그먼트 주입
    matched = enrich_records(new_records, corp_map, seg_map)
    print(f"  세그먼트 매칭 완료: {matched}/{len(new_records)}건")

    # 5. HTML 주입
    print("[4/5] HTML 주입")

    # index.html (13개 필드)
    INDEX_FIELDS = [
        "chatId", "date", "company", "oldTag",
        "primaryTag", "secondaryTag", "confidence",
        "detailCategory", "stabilityStatus", "growthStatus", "limitTier",
        "firstAnswerSec", "closeSec",
    ]

    existing_index, _ = load_existing_records(INDEX_HTML)
    print(f"  index.html 기존: {len(existing_index)}건")
    all_index = existing_index + new_records
    inject_into_html(INDEX_HTML, all_index, INDEX_FIELDS)
    print(f"  index.html 업데이트: {len(all_index)}건")

    # 날짜 업데이트 (헤더의 최종 업데이트 날짜)
    all_dates = [r.get("date") for r in all_index if r.get("date")]
    if all_dates:
        max_date = max(all_dates)
        for html_path in [INDEX_HTML]:
            if os.path.exists(html_path):
                with open(html_path, "r", encoding="utf-8") as f:
                    content = f.read()
                # "2026. 2. 20." 형식으로 변환
                from datetime import datetime
                dt = datetime.strptime(max_date, "%Y-%m-%d")
                date_display = f"{dt.year}. {dt.month}. {dt.day}."
                content = re.sub(
                    r'id="lastUpdateDate">[^<]+<',
                    f'id="lastUpdateDate">{date_display}<',
                    content,
                )
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(content)

    # 요약
    print("\n[5/5] 완료 요약")
    print(f"  ✅ 신규 {len(new_records)}건 추가 (총 {len(all_index)}건)")
    print(f"  ✅ 세그먼트 매칭: {matched}/{len(new_records)}건")
    print(f"  ✅ index.html 업데이트 완료")
    print(f"\n  다음: git commit + push")


if __name__ == "__main__":
    main()
