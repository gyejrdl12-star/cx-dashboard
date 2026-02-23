#!/usr/bin/env python3
"""
CX 대시보드 Phase 1: Excel → JSON 파싱 + 중복 제거

사용법:
  python3 scripts/phase1_parse_excel.py data/_2_19.xlsx

동작:
  1. index.html에서 기존 chatId 추출 (dedup)
  2. Excel 파싱 (UserChat + User data)
  3. 신규 레코드만 JSON 출력 → data/new_records.json

참고: AI 재태깅(primaryTag 등)은 별도 수동 작업. 이 스크립트 범위 밖.
"""

import sys
import os
import json
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

# ── 경로 설정 ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)
INDEX_HTML = os.path.join(ROOT_DIR, "index.html")
OUTPUT_JSON = os.path.join(ROOT_DIR, "data", "new_records.json")


def load_existing_chat_ids():
    """index.html에서 기존 ALL_RECORDS의 chatId 셋 추출"""
    if not os.path.exists(INDEX_HTML):
        print("  ⚠️  index.html 없음 — 전체 신규 처리")
        return set()

    with open(INDEX_HTML, "r", encoding="utf-8") as f:
        content = f.read()

    match = re.search(r"const ALL_RECORDS = (\[.*?\]);", content, re.DOTALL)
    if not match:
        print("  ⚠️  ALL_RECORDS 찾을 수 없음 — 전체 신규 처리")
        return set()

    records = json.loads(match.group(1))
    ids = {r["chatId"] for r in records}
    print(f"  기존 레코드: {len(ids)}건")
    return ids


def build_user_map(wb):
    """User data 시트에서 userId → {company, corp_number} 매핑"""
    ws = wb["User data"]
    user_map = {}
    for r in range(2, ws.max_row + 1):
        uid = ws.cell(r, 1).value
        if not uid:
            continue
        user_map[uid] = {
            "company": ws.cell(r, 5).value,      # profile.companyname
            "corp_number": ws.cell(r, 6).value,   # profile.corp_number
        }
    return user_map


def parse_userchat(wb, user_map, existing_ids):
    """UserChat 시트 파싱 → 신규 레코드 리스트"""
    ws = wb["UserChat"]
    new_records = []
    skipped_dup = 0
    skipped_empty = 0

    for r in range(2, ws.max_row + 1):
        chat_id = ws.cell(r, 1).value
        if not chat_id:
            skipped_empty += 1
            continue

        # 중복 체크
        if chat_id in existing_ids:
            skipped_dup += 1
            continue

        # 날짜
        managed_at = ws.cell(r, 2).value
        if isinstance(managed_at, datetime):
            date_str = managed_at.strftime("%Y-%m-%d")
        elif managed_at:
            date_str = str(managed_at)[:10]
        else:
            date_str = None

        # 법인명: data_only 캐시 → User data 폴백
        company = ws.cell(r, 3).value
        if company and not isinstance(company, str):
            company = None  # formula object
        if not company:
            user_id = ws.cell(r, 14).value
            if user_id and user_id in user_map:
                company = user_map[user_id].get("company")

        # 기존 태그
        old_tag = ws.cell(r, 9).value or "미분류"

        # 응대 시간
        first_answer_sec = ws.cell(r, 67).value  # timeToFirstAnswer (col index 66, 1-based=67)
        close_sec = ws.cell(r, 79).value          # timeToClose (col index 78, 1-based=79)

        # 숫자 변환
        if first_answer_sec is not None:
            try:
                first_answer_sec = float(first_answer_sec)
            except (ValueError, TypeError):
                first_answer_sec = None

        if close_sec is not None:
            try:
                close_sec = float(close_sec)
            except (ValueError, TypeError):
                close_sec = None

        record = {
            "chatId": chat_id,
            "date": date_str,
            "company": company,
            "oldTag": old_tag,
            "primaryTag": old_tag,              # 채널톡 원본 태그 그대로
            "secondaryTag": None,
            "confidence": None,
            "detailCategory": "미매칭",          # ← Phase 2에서 채움
            "stabilityStatus": "미매칭",         # ← Phase 2에서 채움
            "growthStatus": "미매칭",            # ← Phase 2에서 채움
            "limitTier": "미매칭",               # ← Phase 2에서 채움
            "firstAnswerSec": first_answer_sec,
            "closeSec": close_sec,
        }
        new_records.append(record)

    print(f"  빈 행 스킵: {skipped_empty}건")
    print(f"  중복 스킵: {skipped_dup}건")
    return new_records


def main():
    if len(sys.argv) < 2:
        print("사용법: python3 scripts/phase1_parse_excel.py <엑셀파일경로>")
        print("예시:   python3 scripts/phase1_parse_excel.py data/_2_19.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not os.path.isabs(excel_path):
        excel_path = os.path.join(ROOT_DIR, excel_path)

    if not os.path.exists(excel_path):
        print(f"❌ 파일 없음: {excel_path}")
        sys.exit(1)

    print(f"📋 Excel: {os.path.basename(excel_path)}")
    print()

    # 1. 기존 chatId 로드
    print("[1/4] 기존 대시보드 chatId 로드")
    existing_ids = load_existing_chat_ids()

    # 2. Excel 로드
    print("[2/4] Excel 로드")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"  시트: {wb.sheetnames}")

    # 3. 보조 데이터 빌드
    print("[3/3] User 매핑 빌드 + UserChat 파싱")
    user_map = build_user_map(wb)
    print(f"  User 매핑: {len(user_map)}명")

    new_records = parse_userchat(wb, user_map, existing_ids)

    # 결과 출력
    print()
    print(f"✅ 신규 레코드: {len(new_records)}건")

    if not new_records:
        print("   새로운 데이터가 없습니다.")
        return

    # 날짜 분포
    date_counts = {}
    for r in new_records:
        d = r.get("date", "unknown")
        date_counts[d] = date_counts.get(d, 0) + 1
    print(f"   날짜 범위: {min(date_counts.keys())} ~ {max(date_counts.keys())}")
    for d in sorted(date_counts.keys()):
        print(f"     {d}: {date_counts[d]}건")

    # 회사명 커버리지
    with_company = sum(1 for r in new_records if r["company"])
    print(f"   회사명 있음: {with_company}/{len(new_records)}")

    # JSON 저장
    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(new_records, f, ensure_ascii=False, indent=2)

    print(f"\n💾 저장: {os.path.relpath(OUTPUT_JSON, ROOT_DIR)}")
    print(f"   다음 단계: Phase 2 실행 (세그먼트 매칭 + HTML 주입)")


if __name__ == "__main__":
    main()
